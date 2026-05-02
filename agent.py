# ─── مغز ایجنت فروش ─────────────────────────────────────
import os
import openpyxl
from config import CUSTOMERS_FILE, INVENTORY_FILE, MAX_SUGGESTIONS, BUSINESS_NAME


def load_customers():
    if not os.path.exists(CUSTOMERS_FILE):
        return {}
    wb = openpyxl.load_workbook(CUSTOMERS_FILE)
    first_sheet = wb.worksheets[0]

    customers = {}
    for row in first_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            name  = str(row[0]).strip()
            phone = str(row[1]).strip().replace("+", "").replace(" ", "").replace("-", "")
            customers[name] = {"phone": phone, "history": []}

    for sheet in wb.worksheets[1:]:
        cname = sheet.title.strip()
        if cname not in customers:
            continue
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                customers[cname]["history"].append(str(row[0]).strip())

    return customers


def load_inventory():
    if not os.path.exists(INVENTORY_FILE):
        return []
    wb = openpyxl.load_workbook(INVENTORY_FILE)
    sheet = wb.worksheets[0]
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            items.append({
                "name":  str(row[0]).strip(),
                "price": int(row[1]) if row[1] else 0,
            })
    return items


def find_suggestions(history, inventory):
    if not history:
        return inventory[:MAX_SUGGESTIONS]

    keywords = []
    for item in history:
        keywords.extend([w for w in item.split() if len(w) > 2])

    scored = []
    for inv in inventory:
        score = sum(1 for kw in keywords if kw in inv["name"])
        already = any(h.strip() == inv["name"] for h in history)
        if not already:
            scored.append((score, inv))

    scored.sort(key=lambda x: x[0], reverse=True)
    return [item for _, item in scored[:MAX_SUGGESTIONS]]


def build_message(customer_name, history, suggestions):
    msg = f"سلام {customer_name} عزیز 🌹\n\n"
    msg += f"از طرف {BUSINESS_NAME} خدمت شما هستیم.\n\n"

    if history:
        msg += "با توجه به خریدهای قبلی شما، محصولات زیر را ویژه شما پیشنهاد می‌دهیم:\n\n"
    else:
        msg += "جدیدترین محصولات ما برای شما:\n\n"

    for item in suggestions:
        price = f"{item['price']:,}".replace(",", "،")
        msg += f"🔹 {item['name']}\n"
        msg += f"   💰 قیمت: {price} تومان\n\n"

    msg += "برای سفارش یا اطلاعات بیشتر پاسخ دهید 🙏"
    return msg